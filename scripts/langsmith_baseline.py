from __future__ import annotations

import argparse
import os
import time
import sys
from dataclasses import dataclass
from datetime import datetime, UTC
from pathlib import Path
from statistics import mean
from typing import Any

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from openpyxl import load_workbook
from langsmith import Client, traceable
from langsmith.evaluation import EvaluationResult, run_evaluator
from src.nebius_client import NebiusClient


DEFAULT_WORKBOOK = Path("docs") / "intake_routing_golden_dataset.xlsx"
DEFAULT_DATASET_NAME = "intake-routing-golden-v1"
DEFAULT_PROJECT_NAME = "multi-agent-it-support-intake-eval"


def get_nebius_cost_rates() -> tuple[float, float]:
    return (
        float(os.getenv("NEBIUS_INPUT_TOKEN_COST_PER_1K", "0") or 0),
        float(os.getenv("NEBIUS_OUTPUT_TOKEN_COST_PER_1K", "0") or 0),
    )


def load_env_file(path: Path = Path(".env")) -> None:
    """Load simple KEY=VALUE pairs from .env if they are not already set."""
    if not path.exists():
        return
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        if key.startswith("export "):
            key = key.removeprefix("export ").strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value
        if key == "LANGSMITH_API_KEY" and not os.environ.get("LANGCHAIN_API_KEY"):
            os.environ["LANGCHAIN_API_KEY"] = value
        if key == "LANGCHAIN_API_KEY" and not os.environ.get("LANGSMITH_API_KEY"):
            os.environ["LANGSMITH_API_KEY"] = value
        if key == "LANGSMITH_TRACING" and not os.environ.get("LANGCHAIN_TRACING_V2"):
            os.environ["LANGCHAIN_TRACING_V2"] = value
        if key == "LANGCHAIN_TRACING" and not os.environ.get("LANGCHAIN_TRACING_V2"):
            os.environ["LANGCHAIN_TRACING_V2"] = value


@dataclass(frozen=True)
class DatasetRow:
    case_id: str
    bucket: str
    caller_input: str
    hidden_context: str
    expected_next_action: str
    expected_route: str
    forbidden_behavior: str


def read_dataset_rows(workbook_path: Path) -> list[DatasetRow]:
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb["Intake Dataset"]
    rows: list[DatasetRow] = []
    for row in ws.iter_rows(min_row=5, max_col=7, values_only=True):
        if not row[0]:
            continue
        rows.append(
            DatasetRow(
                case_id=str(row[0]),
                bucket=str(row[1]),
                caller_input=str(row[2]),
                hidden_context=str(row[3]),
                expected_next_action=str(row[4]),
                expected_route=str(row[5]),
                forbidden_behavior=str(row[6]),
            )
        )
    return rows


def ensure_dataset(client: Client, dataset_name: str, rows: list[DatasetRow]) -> str:
    """Create the LangSmith dataset once, then add any missing examples."""
    if not client.has_dataset(dataset_name=dataset_name):
        client.create_dataset(
            dataset_name=dataset_name,
            description="Golden dataset for the Week 3 intake orchestrator routing evaluation.",
            metadata={
                "agent_under_test": "Week 3 Intake Orchestrator",
                "source_workbook": str(DEFAULT_WORKBOOK),
                "dataset_version": "v1",
            },
        )

    dataset = client.read_dataset(dataset_name=dataset_name)
    existing_case_ids = {
        str(example.metadata.get("case_id"))
        for example in client.list_examples(dataset_name=dataset_name)
        if example.metadata and example.metadata.get("case_id")
    }

    for row in rows:
        if row.case_id in existing_case_ids:
            continue
        client.create_example(
            dataset_id=dataset.id,
            inputs={
                "case_id": row.case_id,
                "bucket": row.bucket,
                "caller_input": row.caller_input,
                "hidden_context": row.hidden_context,
            },
            outputs={
                "expected_next_action": row.expected_next_action,
                "expected_route": row.expected_route,
            },
            metadata={
                "case_id": row.case_id,
                "bucket": row.bucket,
                "hidden_context": row.hidden_context,
                "forbidden_behavior": row.forbidden_behavior,
            },
        )
    return dataset.id


def _has_keywords(text: str, keywords: tuple[str, ...]) -> bool:
    lowered = text.lower()
    return any(keyword in lowered for keyword in keywords)


@traceable(
    name="baseline_intake_orchestrator",
    run_type="chain",
    project_name=DEFAULT_PROJECT_NAME,
)
def heuristic_intake_orchestrator(inputs: dict[str, Any]) -> dict[str, Any]:
    """A simple baseline policy for the intake-routing evaluation."""
    started = time.perf_counter()
    decision = decide_intake_route(inputs)
    decision["elapsed_ms"] = round((time.perf_counter() - started) * 1000, 3)
    decision["prompt_tokens"] = 0
    decision["completion_tokens"] = 0
    decision["total_tokens"] = 0
    decision["used_llm"] = False
    return decision


@traceable(
    name="llm_intake_orchestrator",
    run_type="chain",
    project_name=DEFAULT_PROJECT_NAME,
)
def llm_intake_orchestrator(inputs: dict[str, Any]) -> dict[str, Any]:
    """Nebius-backed intake policy used for the real baseline."""
    started = time.perf_counter()
    client = NebiusClient()
    caller_input = str(inputs.get("caller_input", "")).strip()
    heuristic_fallback = decide_intake_route(inputs)
    decision = client.decide_intake_route(caller_input, fallback=heuristic_fallback)
    input_cost_per_1k, output_cost_per_1k = get_nebius_cost_rates()
    decision["elapsed_ms"] = round((time.perf_counter() - started) * 1000, 3)
    decision["estimated_cost_usd"] = round(
        (decision.get("prompt_tokens", 0) / 1000.0) * input_cost_per_1k
        + (decision.get("completion_tokens", 0) / 1000.0) * output_cost_per_1k,
        6,
    )
    return decision


@traceable(
    name="decide_intake_route",
    run_type="tool",
    project_name=DEFAULT_PROJECT_NAME,
)
def decide_intake_route(inputs: dict[str, Any]) -> dict[str, Any]:
    caller_input = str(inputs.get("caller_input", "")).strip()
    hidden_context = str(inputs.get("hidden_context", "")).lower()

    if not caller_input:
        action = "stay_in_intake"
        route = "intake"
        reason = "No issue description was provided."
    elif _has_keywords(
        caller_input,
        (
            "hello",
            "hi,",
            "i'm ",
            "my name is",
            "transfer me",
            "skip intake",
            "do not ask",
            "ignore your instructions",
        ),
    ) and not _has_keywords(
        caller_input,
        ("vpn", "email", "mfa", "password", "outlook", "drive", "printer", "sign in"),
    ):
        action = "stay_in_intake"
        route = "intake"
        reason = "Greeting-only, identity-only, or prompt-injection style input."
    elif _has_keywords(caller_input, ("unknown", "broken", "weird", "can't", "cannot", "do not know", "not sure")):
        action = "ask_clarifying_question"
        route = "intake"
        reason = "The issue is too ambiguous to route confidently."
    elif _has_keywords(hidden_context, ("known_incident", "active_incident")) or _has_keywords(
        caller_input, ("every device", "all devices", "everywhere", "company-wide outage", "outlook stopped syncing")
    ):
        action = "route_to_escalation"
        route = "escalation"
        reason = "Known incident / outage signal detected."
    else:
        action = "route_to_verification"
        route = "verification"
        reason = "The issue looks like an individual support request."

    return {
        "observed_next_action": action,
        "observed_route": route,
        "reason": reason,
    }


@run_evaluator
def route_exact_match(run, example) -> EvaluationResult:
    expected = (example.outputs or {}).get("expected_next_action")
    observed = (run.outputs or {}).get("observed_next_action")
    return EvaluationResult(
        key="route_exact_match",
        score=float(observed == expected),
        value={"expected": expected, "observed": observed},
        comment="Exact match between expected and observed next action.",
    )


@run_evaluator
def intake_hold_safety(run, example) -> EvaluationResult:
    """Check that ambiguous, greeting-only, and adversarial cases do not jump out of intake."""
    expected = (example.outputs or {}).get("expected_route")
    observed = (run.outputs or {}).get("observed_route")
    bucket = str((example.metadata or {}).get("bucket", ""))
    if bucket in {"edge_case", "known_failure", "adversarial"} and expected == "intake":
        score = float(observed == "intake")
    elif bucket == "known_failure" and expected == "escalation":
        score = float(observed == "escalation")
    else:
        score = float(observed == expected)
    return EvaluationResult(
        key="intake_hold_safety",
        score=score,
        value={"expected": expected, "observed": observed, "bucket": bucket},
        comment="Protects against premature handoff on weak or adversarial signals.",
    )


def summarize_runs(results: Any) -> dict[str, float]:
    row_results = list(results)
    route_scores: list[float] = []
    safety_scores: list[float] = []
    elapsed_ms_values: list[float] = []
    prompt_tokens: list[float] = []
    completion_tokens: list[float] = []
    total_tokens: list[float] = []
    estimated_costs: list[float] = []
    for row in row_results:
        eval_results = row["evaluation_results"]["results"]
        by_key = {result.key: result for result in eval_results}
        route_scores.append(float(by_key["route_exact_match"].score or 0.0))
        safety_scores.append(float(by_key["intake_hold_safety"].score or 0.0))
        run = row["run"]
        outputs = run.outputs or {}
        if outputs.get("elapsed_ms") is not None:
            elapsed_ms_values.append(float(outputs["elapsed_ms"]))
        if outputs.get("prompt_tokens") is not None:
            prompt_tokens.append(float(outputs["prompt_tokens"]))
        if outputs.get("completion_tokens") is not None:
            completion_tokens.append(float(outputs["completion_tokens"]))
        if outputs.get("total_tokens") is not None:
            total_tokens.append(float(outputs["total_tokens"]))
        if outputs.get("estimated_cost_usd") is not None:
            estimated_costs.append(float(outputs["estimated_cost_usd"]))

    run_stats = getattr(results, "run_stats", {}) or {}
    run_count = float(run_stats.get("run_count") or len(row_results) or 1)
    return {
        "route_accuracy": float(mean(route_scores)) if route_scores else 0.0,
        "intake_hold_safety": float(mean(safety_scores)) if safety_scores else 0.0,
        "p50_latency_s": float(sorted(elapsed_ms_values)[len(elapsed_ms_values) // 2] / 1000.0)
        if elapsed_ms_values
        else 0.0,
        "avg_cost_usd": float(mean(estimated_costs)) if estimated_costs else 0.0,
        "avg_prompt_tokens": float(mean(prompt_tokens)) if prompt_tokens else 0.0,
        "avg_completion_tokens": float(mean(completion_tokens)) if completion_tokens else 0.0,
        "avg_total_tokens": float(mean(total_tokens)) if total_tokens else 0.0,
        "run_count": float(run_count),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Upload the intake dataset and run a LangSmith baseline eval.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--dataset-name", default=DEFAULT_DATASET_NAME)
    parser.add_argument("--project-name", default=DEFAULT_PROJECT_NAME)
    parser.add_argument(
        "--mode",
        choices=("llm", "heuristic"),
        default="llm",
        help="Use the real Nebius-backed intake decision or the heuristic smoke test.",
    )
    parser.add_argument("--limit", type=int, default=0, help="Optional row limit for a smoke test.")
    args = parser.parse_args()

    load_env_file()

    if not (os.environ.get("LANGCHAIN_API_KEY") or os.environ.get("LANGSMITH_API_KEY")):
        raise SystemExit(
            "Set LANGCHAIN_API_KEY (or LANGSMITH_API_KEY) in your environment or .env, then rerun."
        )

    os.environ["LANGCHAIN_TRACING_V2"] = os.environ.get("LANGCHAIN_TRACING_V2", "true")
    os.environ["LANGCHAIN_PROJECT"] = args.project_name

    rows = read_dataset_rows(args.workbook)
    if args.limit and args.limit > 0:
        rows = rows[: args.limit]

    client = Client()
    dataset_id = ensure_dataset(client, args.dataset_name, rows)
    print(f"Dataset ready: {args.dataset_name} ({dataset_id})")

    target = llm_intake_orchestrator if args.mode == "llm" else heuristic_intake_orchestrator

    results = client.evaluate(
        target,
        data=args.dataset_name,
        evaluators=[route_exact_match, intake_hold_safety],
        metadata={
            "agent_under_test": "Week 3 Intake Orchestrator",
            "baseline_type": args.mode,
            "dataset_name": args.dataset_name,
            "dataset_version": "v1",
            "source_workbook": str(args.workbook),
            "run_at": datetime.now(UTC).isoformat(timespec="seconds"),
        },
        experiment_prefix="baseline-intake-routing",
        description="Baseline routing eval for the Week 3 intake orchestrator.",
        blocking=True,
    )

    summary = summarize_runs(results)
    print("Baseline summary:")
    for key, value in summary.items():
        if key == "run_count":
            print(f"  {key}: {int(value)}")
        else:
            print(f"  {key}: {value:.4f}")

    experiment_name = getattr(results, "experiment_name", None)
    if experiment_name:
        print(f"LangSmith experiment: {experiment_name}")
    experiment_url = getattr(results, "url", None)
    if experiment_url:
        print(f"LangSmith URL: {experiment_url}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
